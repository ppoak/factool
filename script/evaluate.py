from pathlib import Path
from typing import Dict, Optional, List, Union, Literal

import numpy as np
import pandas as pd
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.legend import Legend

import quool
from factool import Evaluator


def _resolve_col(
    ws, spec: Union[int, str, None], headers: bool, header_row: int, max_col: int
) -> Optional[int]:
    """Resolve a column spec (int index or header name) to a column index."""
    if spec is None:
        return None
    if isinstance(spec, int):
        return spec
    if isinstance(spec, str):
        if not headers:
            raise ValueError(
                "When headers=False, header names cannot be used to select columns."
            )
        for c in range(1, max_col + 1):
            if ws.cell(row=header_row, column=c).value == spec:
                return c
        raise ValueError(f"Header column not found: {spec}")
    raise ValueError("Column spec must be an integer index or header string.")


def _add_line_chart(
    ws,
    title: Optional[str] = None,
    anchor: str = "G2",
    headers: bool = True,
    category_col: Optional[
        Union[int, str]
    ] = 1,  # Category axis column (default: 1st column)
    series_cols: Optional[
        List[Union[int, str]]
    ] = None,  # Data series columns (default: 2..last)
    x_axis: bool = True,
    y_axis: bool = True,
    legend: Optional[Legend] = None,
    colors: Optional[List[str]] = None,  # Hex RGB like ["FF0000", "00FF00"]
    width: Optional[float] = None,
    height: Optional[float] = None,
):
    """Add a line chart to the worksheet."""
    if ws.max_row < 2 or ws.max_column < 2:
        return None

    max_row = ws.max_row
    max_col = ws.max_column
    header_row = 1
    data_start_row = 2 if headers else 1

    chart = LineChart()
    chart.title = title or ""
    if width is not None:
        chart.width = width
    if height is not None:
        chart.height = height

    cat_col = _resolve_col(ws, category_col, headers, header_row, max_col)

    if series_cols is None:
        series_cols = [c for c in range(2, max_col + 1) if c != cat_col]

    # Add each selected column as a series. Use titles_from_data=headers for names.
    for col_spec in series_cols:
        col = _resolve_col(ws, col_spec, headers, header_row, max_col)
        data_ref = Reference(
            ws,
            min_col=col,
            min_row=header_row if headers else data_start_row,
            max_col=col,
            max_row=max_row,
        )
        chart.add_data(data_ref, titles_from_data=headers)

    if cat_col is None:
        raise ValueError("Line chart requires category_col.")

    cats = Reference(ws, min_col=cat_col, min_row=data_start_row, max_row=max_row)
    chart.set_categories(cats)

    # Apply colors if provided
    if colors:
        for i, s in enumerate(chart.series):
            if i < len(colors):
                s.graphicalProperties.line.solidFill = colors[i]
    chart.x_axis.delete = not x_axis
    chart.y_axis.delete = not y_axis
    chart.legend = legend

    ws.add_chart(chart, anchor)
    return chart


def _add_bar_chart(
    ws,
    title: Optional[str] = None,
    anchor: str = "G2",
    headers: bool = True,
    category_col: Optional[
        Union[int, str]
    ] = 1,  # Category axis column (default: 1st column)
    series_cols: Optional[
        List[Union[int, str]]
    ] = None,  # Data series columns (default: 2..last)
    x_axis: bool = True,
    y_axis: bool = True,
    legend: Optional[Legend] = None,
    colors: Optional[List[str]] = None,  # Hex RGB like ["FF0000", "00FF00"]
    width: Optional[float] = None,
    height: Optional[float] = None,
    bar_type: str = "col",  # "col" (vertical) or "bar" (horizontal)
    grouping: Optional[str] = None,  # None | "clustered" | "stacked" | "percentStacked"
):
    """Add a bar chart to the worksheet."""
    if ws.max_row < 2 or ws.max_column < 2:
        return None

    max_row = ws.max_row
    max_col = ws.max_column
    header_row = 1
    data_start_row = 2 if headers else 1

    chart = BarChart()
    chart.title = title or ""
    chart.type = bar_type
    if grouping:
        chart.grouping = grouping
    if width is not None:
        chart.width = width
    if height is not None:
        chart.height = height

    cat_col = _resolve_col(ws, category_col, headers, header_row, max_col)

    if series_cols is None:
        series_cols = [c for c in range(2, max_col + 1) if c != cat_col]

    # Add each selected column as a series with optional titles from header
    for col_spec in series_cols:
        col = _resolve_col(ws, col_spec, headers, header_row, max_col)
        data_ref = Reference(
            ws,
            min_col=col,
            min_row=header_row if headers else data_start_row,
            max_col=col,
            max_row=max_row,
        )
        chart.add_data(data_ref, titles_from_data=headers)

    if cat_col is None:
        raise ValueError("Bar chart requires category_col.")
    cats = Reference(ws, min_col=cat_col, min_row=data_start_row, max_row=max_row)
    chart.set_categories(cats)

    # Apply fill colors for bars
    if colors:
        for i, s in enumerate(chart.series):
            if i < len(colors):
                s.graphicalProperties.solidFill = colors[i]

    chart.x_axis.delete = not x_axis
    chart.y_axis.delete = not y_axis
    chart.legend = legend

    ws.add_chart(chart, anchor)
    return chart


def _add_scatter_chart(
    ws,
    title: Optional[str] = None,
    anchor: str = "G2",
    headers: bool = True,
    x_col: Optional[
        Union[int, str]
    ] = None,  # X axis column (defaults to category_col if set)
    y_cols: Optional[
        List[Union[int, str]]
    ] = None,  # Y series columns (default: all except X)
    category_col: Optional[
        Union[int, str]
    ] = 1,  # Fallback for X axis when x_col is None
    x_axis: bool = True,
    y_axis: bool = True,
    legend: Optional[Legend] = None,
    colors: Optional[List[str]] = None,  # Hex RGB like ["FF0000", "00FF00"]
    width: Optional[float] = None,
    height: Optional[float] = None,
    scatter_style: Literal[
        "marker", "line", "lineMarker"
    ] = "marker",  # "marker" | "line" | "lineMarker"
):
    """Add a scatter chart to the worksheet."""
    if ws.max_row < 2 or ws.max_column < 2:
        return None

    max_row = ws.max_row
    max_col = ws.max_column
    header_row = 1
    data_start_row = 2 if headers else 1

    chart = ScatterChart()
    chart.title = title or ""
    chart.scatterStyle = scatter_style
    if width is not None:
        chart.width = width
    if height is not None:
        chart.height = height

    x_c = (
        _resolve_col(ws, x_col, headers, header_row, max_col)
        if x_col is not None
        else _resolve_col(ws, category_col, headers, header_row, max_col)
    )
    if x_c is None:
        raise ValueError("Scatter chart requires x_col or category_col as X axis.")
    xvalues = Reference(ws, min_col=x_c, min_row=data_start_row, max_row=max_row)

    if y_cols is None:
        y_cols = [c for c in range(1, max_col + 1) if c != x_c]

    for idx, col_spec in enumerate(y_cols):
        col = _resolve_col(ws, col_spec, headers, header_row, max_col)
        yvalues = Reference(ws, min_col=col, min_row=data_start_row, max_row=max_row)
        s = Series(yvalues, xvalues)
        s.marker.symbol = "circle"
        s.marker.size = 5
        s.graphicalProperties.line.noFill = True
        chart.series.append(s)
        if colors and idx < len(colors):
            s.marker.graphicalProperties.solidFill = colors[idx]
            s.marker.graphicalProperties.line.noFill = True

    chart.x_axis.delete = not x_axis
    chart.y_axis.delete = not y_axis
    chart.legend = legend

    ws.add_chart(chart, anchor)
    return chart


def run_factor_pipeline(
    factor: pd.DataFrame,
    price: pd.DataFrame,
    feasible: Optional[pd.DataFrame] = None,
    weight: Optional[pd.DataFrame] = None,
    # Grouping/HL params
    n_groups: int = 10,
    horizon: int = 1,
    skip_horizon: bool = True,
    bucketing_mode: str = "conditional",
    # IC params
    ic_method: str = "spearman",
    # Rolling TS exposure params
    ts_window: int = 252,
    ts_min_obs: int = 60,
    ts_intercept: bool = True,
    ts_n_jobs: int = -1,
    # Cross-sectional regression params
    cs_add_intercept: bool = True,
    cs_cov_type: str = "white",
    cs_white_type: str = "HC1",
    # Fama-MacBeth params
    fmb_nw_lag: int = 3,
    # GMM pricing params
    run_gmm: bool = True,
) -> Dict[str, object]:
    """Run a full factor evaluation pipeline and return computed results.

    This function orchestrates a standard empirical asset-pricing workflow using a single
    factor (or a list of factors) and price data via factool.Evaluator. It computes:
    - Information coefficients (IC) and their mean/t-stats
    - Grouped (bucketed) portfolio returns and cumulative values, plus evaluation metrics
    - Rolling time-series exposure of the factor to the grouped long-short (HL) portfolio
    - Cross-sectional regressions of returns on the factor
    - Fama–MacBeth estimates and t-stats
    - GRS test statistics for joint alpha = 0
    - Optional GMM linear pricing test

    Args:
        factor (pd.DataFrame or List[pd.DataFrame]): Factor values. Typically a DataFrame
            (or a list of DataFrames) indexed by a MultiIndex with date at level 0 and
            asset identifiers (e.g., codes) at level 1. Values should be numeric and aligned
            with the price DataFrame.
        price (pd.DataFrame): Price or return data used by the evaluator. Must be aligned
            with the factor on index and identifiers as required by factool.Evaluator.
        feasible (Optional[pd.DataFrame]): Optional boolean mask indicating tradable/feasible
            assets by date and code. Aligns with factor/price indices.
        weight (Optional[pd.DataFrame]): Optional cross-sectional weights by date and code
            for portfolio construction and regressions.
        n_groups (int): Number of buckets used to sort assets by the factor for grouped
            portfolio analysis. Default is 10.
        horizon (int): Return horizon (in periods) for IC and performance calculations.
            Default is 1.
        skip_horizon (bool): Whether to skip the immediate next horizon when forming future
            returns (e.g., to avoid look-ahead overlap). Default is True.
        bucketing_mode (str): Bucketing mode for grouped portfolios. Common options include
            "conditional" and "single". Default is "conditional".
        ic_method (str): Method for information coefficient (e.g., "spearman"). Default is
            "spearman".
        ts_window (int): Rolling window length for time-series exposure estimation. Default
            is 252.
        ts_min_obs (int): Minimum number of observations required within the rolling window
            to fit the time-series regression. Default is 60.
        ts_intercept (bool): Whether to include an intercept in the time-series regression.
            Default is True.
        ts_n_jobs (int): Number of parallel jobs for time-series exposure computation.
            Use -1 to utilize all available cores. Default is -1.
        cs_add_intercept (bool): Whether to add an intercept in cross-sectional regressions.
            Default is True.
        cs_cov_type (str): Covariance type for cross-sectional regression (e.g., "white").
            Default is "white".
        cs_white_type (str): White heteroskedasticity-consistent estimator type (e.g., "HC1").
            Default is "HC1".
        fmb_nw_lag (int): Newey–West lag length for Fama–MacBeth t-stat computation.
            Default is 3.
        run_gmm (bool): Whether to run the GMM linear pricing test. Default is True.

    Returns:
        Dict[str, object]: A dictionary containing:
            - "ic": Information coefficients by date (pd.Series or pd.DataFrame).
            - "ic_mean": Mean IC by factor/column (pd.Series).
            - "ic_tstat": IC t-statistics (pd.Series).
            - "factor_return": Sorted HL factor portfolio returns by date (pd.Series).
            - "group_returns": Grouped portfolio returns (pd.DataFrame), including HL.
            - "group_value": Cumulative value of grouped portfolios (pd.DataFrame).
            - "group_eval": Evaluation metrics for each grouped portfolio column (pd.DataFrame).
            - "factor_exposure_mean": Mean time-series exposure by date (pd.DataFrame).
            - "factor_exposure_mean_t": Mean t-stats for exposures by date (pd.DataFrame).
            - "factor_premia": Cross-sectional factor premia by date (pd.Series).
            - "factor_premia_t": Cross-sectional t-stats by date (pd.Series).
            - "r2": Cross-sectional regression R-squared by date (pd.Series).
            - "fmb_premia": Fama–MacBeth premia (pd.Series).
            - "fmb_premia_t": Fama–MacBeth t-stats (pd.Series).
            - "grs_stat": GRS test statistic (float).
            - "grs_pval": GRS test p-value (float).
            - "gmm_result": Dict with GMM results (e.g., {"J": float, "pval": float}) or None.

    Raises:
        ValueError: If input indices are misaligned or parameters are invalid for the evaluator.
        RuntimeError: If underlying evaluator methods fail.
        Exception: Any error propagated from factool.Evaluator, quool, or NumPy/pandas.

    Notes:
        - The code groups by level=0, assuming the first index level is the date.
        - group_eval is computed via quool.Evaluator.evaluate applied column-wise to group_value.
    """
    e = Evaluator(factor=factor, price=price)
    asset_returns = e._future_return(horizon=horizon, skip=skip_horizon)

    # Step 1: IC and direction
    e.get_info_coef(horizon=horizon, skip_horizon=skip_horizon, method=ic_method)
    ic = e.ic
    ic_mean = e.ic.mean()
    ic_tstat = ic_mean / (e.ic.std() / np.sqrt(e.ic.shape[0]))

    # Step 2: Grouped portfolios and HL
    e.get_group_returns(
        n=n_groups,
        horizon=horizon,
        skip_horizon=skip_horizon,
        mode=bucketing_mode,
        feasible=feasible,
        weight=weight,
    )
    factor_return = e.sorted_factor_return
    group_returns = pd.concat(
        [gr.groupby(level=0).mean() for gr in e.group_returns.values()]
        + [factor_return],
        axis=1,
    )
    group_returns_mean = group_returns.mean()
    group_returns_t = group_returns_mean / (
        group_returns.std() / np.sqrt(group_returns.shape[0])
    )

    group_value = (
        1 + group_returns.shift(1).dropna(how="all", axis=0).fillna(0)
    ).cumprod()
    group_eval = group_value.apply(quool.Evaluator.evaluate)

    # Step 3: Rolling TS exposure vs HL
    e.get_factor_exposure(
        horizon=horizon,
        feasible=feasible,
        window=ts_window,
        min_obs=ts_min_obs,
        intercept=ts_intercept,
        n_jobs=ts_n_jobs,
    )
    factor_exposure_mean = e.factor_exposure.groupby(level=0).mean()
    factor_exposure_mean_t = e.factor_exposure_t.groupby(level=0).mean()

    # Step 4: Cross-sectional regression on the single factor
    e.cross_sectional_regression(
        horizon=horizon,
        feasible=feasible,
        weight=weight,
        add_intercept=cs_add_intercept,
        cov_type=cs_cov_type,
        white_type=cs_white_type,
    )
    factor_premia = e.factor_premia
    factor_premia_t = e.factor_premia_t
    factor_r2 = e.factor_r2

    # Step 5: Fama-MacBeth regression
    e.fama_macbeth(nw_lag=fmb_nw_lag)
    fmb_premia = e.fmb_premia
    fmb_premia_t = e.fmb_tstats

    # Step 6: GRS test (joint alpha = 0)
    e.grs_test(horizon=horizon, add_intercept=True)
    grs_stat, grs_pval = e.grs_stat, e.grs_pval

    # Step 7: GMM pricing (optional) with HL factor
    gmm_result = None
    if run_gmm:
        e.gmm_linear_pricing(horizon=horizon, two_step=True)
        gmm_result = e.gmm_result

    return {
        "asset_returns": asset_returns,
        "ic": ic,
        "ic_mean": ic_mean,
        "ic_tstat": ic_tstat,
        "factor_return": factor_return,
        "group_returns": group_returns,
        "group_returns_mean": group_returns_mean,
        "group_returns_t": group_returns_t,
        "group_value": group_value,
        "group_eval": group_eval,
        "factor_exposure_mean": factor_exposure_mean,
        "factor_exposure_mean_t": factor_exposure_mean_t,
        "factor_premia": factor_premia,
        "factor_premia_t": factor_premia_t,
        "r2": factor_r2,
        "fmb_premia": fmb_premia,
        "fmb_premia_t": fmb_premia_t,
        "grs_stat": grs_stat,
        "grs_pval": grs_pval,
        "gmm_result": gmm_result,
    }


def save_factor_pipeline(
    factor: Union[pd.DataFrame, List[pd.DataFrame]],
    price: pd.DataFrame,
    output_path: Union[str, Path],
    feasible: Optional[pd.DataFrame] = None,
    weight: Optional[pd.DataFrame] = None,
    n_groups: int = 10,
    horizon: int = 1,
    skip_horizon: bool = True,
    bucketing_mode: str = "conditional",
    ic_method: str = "spearman",
    ts_window: int = 252,
    ts_min_obs: int = 60,
    ts_intercept: bool = True,
    ts_n_jobs: int = -1,
    cs_add_intercept: bool = True,
    cs_cov_type: str = "white",
    cs_white_type: str = "HC1",
    fmb_nw_lag: int = 3,
    run_gmm: bool = True,
) -> Dict[str, object]:
    """Run the factor pipeline and save results to an Excel file with multiple sheets.

    This function executes run_factor_pipeline and writes the outputs to an Excel workbook
    with the following sheets:
    - "IC": Daily information coefficients.
    - "Group Return": Average grouped portfolio returns by date (including the HL column).
    - "Group Value": Cumulative values of grouped portfolios.
    - "Group Eval": Evaluation metrics for grouped portfolios.
    - "Time Series Regression": Mean time-series exposures and t-stats by code.
    - "Cross Sectional Regression": Daily factor premia, t-stats, and R-squared.
    - "Stats": Summary statistics including IC means/t-stats, GRS test, Fama–MacBeth
    premia/t-stats, and GMM test results (J and p-value).

    Args:
        factor (Union[pd.DataFrame, List[pd.DataFrame]]): Factor values as a DataFrame or a
            list of DataFrames. Typically indexed by date (level 0) and asset code (level 1).
            Must be compatible with factool.Evaluator.
        price (pd.DataFrame): Price/return data aligned with the factor inputs.
        output_path (Union[str, Path]): Path to the output Excel file (.xlsx). Parent
            directories should exist or be creatable.
        feasible (Optional[pd.DataFrame]): Optional feasibility mask (boolean) aligned to
            factor/price indices.
        weight (Optional[pd.DataFrame]): Optional cross-sectional weights aligned to indices.
        n_groups (int): Number of buckets for grouped portfolios. Default is 10.
        horizon (int): Return horizon in periods. Default is 1.
        skip_horizon (bool): Whether to skip the next horizon window in future returns.
            Default is True.
        bucketing_mode (str): Bucketing mode ("conditional", "single", etc.). Default is
            "conditional".
        ic_method (str): IC computation method (e.g., "spearman"). Default is "spearman".
        ts_window (int): Rolling window for time-series exposures. Default is 252.
        ts_min_obs (int): Minimum observations for rolling regression. Default is 60.
        ts_intercept (bool): Include intercept in time-series regression. Default is True.
        ts_n_jobs (int): Parallel jobs for exposure computation (-1 for all cores). Default
            is -1.
        cs_add_intercept (bool): Include intercept in cross-sectional regression. Default is
            True.
        cs_cov_type (str): Covariance type for cross-sectional regression. Default is "white".
        cs_white_type (str): White estimator type (e.g., "HC1"). Default is "HC1".
        fmb_nw_lag (int): Newey–West lag length for FMB t-stats. Default is 3.
        run_gmm (bool): Whether to run the GMM pricing test and include its results in "Stats".
            Default is True.

    Returns:
        str: A status message including a Markdown-style link to the saved Excel file.

    Raises:
        OSError: If writing the Excel file fails (e.g., path not writable).
        ValueError: If pipeline inputs are invalid or misaligned.
        KeyError or TypeError: If "Stats" sheet assembly fails due to missing GMM results
            (e.g., run_gmm=False yields gmm_result=None).

    Notes:
        - The Excel writer uses the openpyxl engine; ensure the .xlsx extension is used.
        - When run_gmm=False, the current implementation will attempt to access gmm_result
        and may raise an error; set run_gmm=True to include GMM statistics in "Stats".
        - Index names are set when resetting indices to produce clean tabular outputs.
    """
    results = run_factor_pipeline(
        factor=factor,
        price=price,
        feasible=feasible,
        weight=weight,
        n_groups=n_groups,
        horizon=horizon,
        skip_horizon=skip_horizon,
        bucketing_mode=bucketing_mode,
        ic_method=ic_method,
        ts_window=ts_window,
        ts_min_obs=ts_min_obs,
        ts_intercept=ts_intercept,
        ts_n_jobs=ts_n_jobs,
        cs_add_intercept=cs_add_intercept,
        cs_cov_type=cs_cov_type,
        cs_white_type=cs_white_type,
        fmb_nw_lag=fmb_nw_lag,
        run_gmm=run_gmm,
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results["ic"].reset_index().to_excel(writer, index=False, sheet_name="IC")
        pd.concat(
            [
                results["group_returns_mean"].to_frame("Group Return Mean"),
                results["group_returns_t"].to_frame("Group Return T"),
            ],
            axis=1,
        ).reset_index(names=["category"]).to_excel(
            writer, index=False, sheet_name="Group Return"
        )
        results["group_value"].reset_index().to_excel(
            writer, index=False, sheet_name="Group Value"
        )
        results["group_eval"].reset_index(names=["indicator"]).to_excel(
            writer, index=False, sheet_name="Group Eval"
        )

        pd.concat(
            [
                results["factor_exposure_mean"],
                results["factor_exposure_mean_t"],
                results["asset_returns"].mean().to_frame("Asset Return Mean"),
            ],
            axis=1,
        ).reset_index(names=["code"]).to_excel(
            writer, index=False, sheet_name="Time Series Regression"
        )

        pd.concat(
            [
                results["factor_premia"],
                results["factor_premia_t"],
                results["r2"].to_frame(),
            ],
            axis=1,
        ).reset_index(names=["date"]).to_excel(
            writer, index=False, sheet_name="Cross Sectional Regression"
        )

        pd.concat(
            (
                [
                    results["ic_mean"].add_suffix("-ic-mean"),
                    results["ic_tstat"].add_suffix("-ic-t"),
                    pd.Series(
                        {
                            "grs_stat": results["grs_stat"],
                            "grs_pval": results["grs_pval"],
                        },
                    ),
                    results["fmb_premia"].add_suffix("-fmb_premia"),
                    results["fmb_premia_t"].add_suffix("-fmb_premia-t"),
                ]
                + (
                    [
                        pd.Series(
                            {
                                "gmm_J": results["gmm_result"]["J"],
                                "gmm_pval": results["gmm_result"]["pval"],
                            }
                        )
                    ]
                    if run_gmm
                    else []
                )
            ),
        ).to_frame("Stats").reset_index(names=["stat_name"]).to_excel(
            writer, index=False, sheet_name="Stats"
        )

        # IC (line chart over time)
        ws = writer.sheets["IC"]
        _add_bar_chart(
            ws,
            category_col=1,
            title="Information Coefficient",
        )

        # Group Return (bar chart of group returns)
        ws = writer.sheets["Group Return"]
        _add_bar_chart(ws, series_cols=[2], title="Group Portfolio Returns")

        # Group Value (line chart of cumulative values)
        ws = writer.sheets["Group Value"]
        _add_line_chart(ws, title="Group Portfolio Cumulative Value")

        # Time Series Regression (scatter chart of exposures)
        ws = writer.sheets["Time Series Regression"]
        _add_scatter_chart(
            ws,
            x_col=list(range(3, 3 + (len(factor) if isinstance(factor, list) else 1))),
            y_cols=[6],
            title="Time-Series Factor Exposure vs Asset Return",
        )

        # Cross Sectional Regression (bar chart of premia)
        ws = writer.sheets["Cross Sectional Regression"]
        _add_bar_chart(ws, series_cols=[3], title="Cross-Sectional Factor Premia")

    return f"Factor evaluation ended\n\n![result]({str(output_path)})"


if __name__ == "__main__":
    import factool
    import os
    import dotenv

    import parquool

    dotenv.load_dotenv()

    begin = "2015-01-01"
    end = "2025-06-30"
    horizon = 21
    skip_horizon = True
    output_path = "out/debt_to_asset.xlsx"
    n_groups = 10
    bucketing_mode = "single"
    ts_n_jobs = -1
    run_gmm = True

    dps = factool.DuckParquetSource(f"data/barra_bookprice")
    df = dps.get_factor("barra_bookprice", begin=begin, end=end)
    source = factool.DuckParquetSource(os.getenv("QUOTESDAY_PATH"))
    price = source.get_factor("close_post", begin=begin, end=end)

    notifier = parquool.notify_task()
    notifier(save_factor_pipeline)(
        factor=[df],
        price=price,
        horizon=horizon,
        skip_horizon=skip_horizon,
        output_path=output_path,
        n_groups=n_groups,
        bucketing_mode=bucketing_mode,
        ts_n_jobs=ts_n_jobs,
        run_gmm=run_gmm,
    )
