import os
import dotenv
from pathlib import Path
from logging import Logger
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple, Union, Optional, Literal

import numpy as np
import pandas as pd

from openpyxl.utils import get_column_letter
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference

import quool
from parquool import setup_logger

from factool import DuckPQSource, Evaluator

import statsmodels.api as sm
from statsmodels.stats.diagnostic import acorr_ljungbox
from scipy import stats

dotenv.load_dotenv()


FactorPath = str


def safe_float(x: Any) -> float:
    try:
        if x is None:
            return np.nan
        if isinstance(x, (np.floating, float, int, np.integer)):
            return float(x)
        return float(np.asarray(x))
    except Exception:
        return np.nan


def get_feasible(
    data_source: DuckPQSource, begin: str, end: str, min_list_days: int = 90
):
    sql = f"""
    SELECT
        q.date AS date,
        q.code AS code,
        (
            q.high > q.limit_down
            AND q.low  < q.limit_up
            AND COALESCE(q.st, false) = false
            AND COALESCE(q.suspended, false) = false
            AND datediff('day', i.listed_date, q.date) > {min_list_days}
        ) AS feasible
    FROM quotes_day AS q
    JOIN instruments_info AS i
        ON q.code = i.code
    WHERE q.date >= '{begin}' AND q.date <= '{end}'
    """

    data = data_source.query(sql)
    data["date"] = pd.to_datetime(data["date"])

    feasible = data.set_index(["date", "code"]).sort_index()
    return feasible


@dataclass
class BacktestParams:
    factor_paths: List[str]
    begin: str
    end: str
    target_path: str = "target/open"
    horizon: int = 5
    baseline_factors: Optional[Union[str, List[str]]] = None

    # Tradability mask and weight
    min_list_days: int = 90
    weight_path: str = None

    # IC and grouping
    ic_method: str = "spearman"
    n_groups: int = 10
    bucketing_mode: str = "single"

    # Cross-sectional regression
    cs_add_intercept: bool = True
    cs_cov_type: str = "white"
    cs_white_type: str = "HC1"

    # IC stability settings
    ic_roll_window: int = 60
    ic_acf_lags: int = 10
    ic_break_k: int = (
        1  # number of candidate breakpoints to test (simple single-break scan)
    )

    # Monotonicity settings
    monotonicity_use_excess: bool = (
        False  # if True, use group returns minus cross-sectional mean
    )


def grenerate_test_key(param: BacktestParams) -> str:
    paths = (
        param.factor_paths
        if isinstance(param.factor_paths, list)
        else [param.factor_paths]
    )
    bases = []
    if param.baseline_factors is not None:
        bases = (
            param.baseline_factors
            if isinstance(param.baseline_factors, list)
            else [param.baseline_factors]
        )
    return "+".join([path.split("/", 1)[-1] for path in paths]) + (
        ""
        if not bases
        else (f"__" + "+".join([base.split("/", 1)[-1] for base in bases]))
    )


def newey_west_tstat(x: pd.Series, lags: Optional[int] = None) -> float:
    """
    Newey-West t-stat for mean(x) with HAC covariance.
    Uses statsmodels OLS of x on constant.
    """
    s = pd.Series(x).dropna()
    if s.shape[0] < 10:
        return np.nan
    y = s.values
    X = np.ones((len(y), 1))
    model = sm.OLS(y, X)
    if lags is None:
        # Common rule-of-thumb
        lags = int(np.floor(4 * (len(y) / 100) ** (2 / 9)))
    try:
        fit = model.fit(cov_type="HAC", cov_kwds={"maxlags": int(lags)})
        return float(fit.tvalues[0])
    except Exception:
        return np.nan


def ic_summary_stats(ic: pd.Series) -> Dict[str, Any]:
    """
    IC summary including win rate, ICIR, tails, and HAC t-stat.
    """
    s = pd.Series(ic).dropna()
    n = int(s.shape[0])
    if n == 0:
        return {
            "ic_n": 0,
            "ic_mean": np.nan,
            "ic_std": np.nan,
            "ic_ir": np.nan,
            "ic_t": np.nan,
            "ic_t_nw": np.nan,
            "ic_win_rate": np.nan,
            "ic_p05": np.nan,
            "ic_p50": np.nan,
            "ic_p95": np.nan,
            "ic_skew": np.nan,
            "ic_kurt": np.nan,
            "ic_tail_gt_2std": np.nan,
        }

    mu = float(s.mean())
    sd = float(s.std(ddof=1)) if n > 1 else np.nan
    ir = mu / sd if sd and np.isfinite(sd) and sd > 0 else np.nan
    t = mu / (sd / np.sqrt(n)) if sd and np.isfinite(sd) and sd > 0 else np.nan
    t_nw = newey_west_tstat(s)

    win_rate = float((s > 0).mean())
    p05, p50, p95 = [float(v) for v in np.nanpercentile(s.values, [5, 50, 95])]
    skew = float(stats.skew(s.values, nan_policy="omit")) if n >= 3 else np.nan
    kurt = (
        float(stats.kurtosis(s.values, fisher=True, nan_policy="omit"))
        if n >= 4
        else np.nan
    )
    tail_gt_2std = float((np.abs(s - mu) > 2 * sd).mean()) if sd and sd > 0 else np.nan

    return {
        "ic_n": n,
        "ic_mean": mu,
        "ic_std": sd,
        "ic_ir": ir,
        "ic_t": t,
        "ic_t_nw": t_nw,
        "ic_win_rate": win_rate,
        "ic_p05": p05,
        "ic_p50": p50,
        "ic_p95": p95,
        "ic_skew": skew,
        "ic_kurt": kurt,
        "ic_tail_gt_2std": tail_gt_2std,
    }


def ic_stability_tests(
    ic: pd.Series, roll_window: int, acf_lags: int, break_k: int = 1
) -> Dict[str, Any]:
    """
    Stability/drift tests:
    - Rolling mean/IR variability
    - ACF(1) and Ljung-Box p-value (autocorrelation)
    - Simple single-break scan (max abs difference in pre/post mean), report best split and p-value
    """
    s = pd.Series(ic).dropna()
    out: Dict[str, Any] = {}
    if s.shape[0] < max(50, roll_window + 5):
        out.update(
            {
                "ic_roll_mean_std": np.nan,
                "ic_roll_ir_std": np.nan,
                "ic_acf1": np.nan,
                "ic_lb_pvalue": np.nan,
                "ic_best_break_date": None,
                "ic_break_stat": np.nan,
                "ic_break_pvalue": np.nan,
            }
        )
        return out

    roll_mean = s.rolling(roll_window).mean()
    roll_std = s.rolling(roll_window).std(ddof=1)
    roll_ir = roll_mean / roll_std

    out["ic_roll_mean_std"] = safe_float(roll_mean.dropna().std(ddof=1))
    out["ic_roll_ir_std"] = safe_float(roll_ir.dropna().std(ddof=1))

    # Autocorrelation tests
    try:
        out["ic_acf1"] = safe_float(s.autocorr(lag=1))
    except Exception:
        out["ic_acf1"] = np.nan

    try:
        lb = acorr_ljungbox(s.values, lags=[min(acf_lags, len(s) - 1)], return_df=True)
        out["ic_lb_pvalue"] = safe_float(lb["lb_pvalue"].iloc[-1])
    except Exception:
        out["ic_lb_pvalue"] = np.nan

    # Simple best single break scan (choose split that maximizes |mean1 - mean2|)
    idx = s.index
    values = s.values
    n = len(s)
    min_seg = max(30, n // 10)
    best = {"stat": -np.inf, "split": None, "pvalue": np.nan}

    # Candidate splits
    split_candidates = list(range(min_seg, n - min_seg))
    if break_k is not None and break_k > 0 and len(split_candidates) > 0:
        # To limit compute, subsample candidates
        step = max(1, len(split_candidates) // 200)
        split_candidates = split_candidates[::step]

    for k in split_candidates:
        a = values[:k]
        b = values[k:]
        if len(a) < min_seg or len(b) < min_seg:
            continue
        # Two-sample t-test (Welch)
        stat, pval = stats.ttest_ind(a, b, equal_var=False, nan_policy="omit")
        diff = np.nanmean(a) - np.nanmean(b)
        score = np.abs(diff)
        if np.isfinite(score) and score > best["stat"]:
            best = {"stat": score, "split": k, "pvalue": pval}

    if best["split"] is None:
        out["ic_best_break_date"] = None
        out["ic_break_stat"] = np.nan
        out["ic_break_pvalue"] = np.nan
    else:
        out["ic_best_break_date"] = str(idx[best["split"]].date())
        out["ic_break_stat"] = safe_float(best["stat"])
        out["ic_break_pvalue"] = safe_float(best["pvalue"])

    return out


def _extract_factor_group_columns(
    group_returns: pd.DataFrame, factor_name: str, n_groups: int
) -> List[str]:
    return [f"{factor_name}({i + 1})" for i in range(n_groups)]


def group_monotonicity_tests(
    e: Evaluator,
    factor_name: str,
    n_groups: int,
    use_excess: bool = False,
) -> Dict[str, Any]:
    """
    Uses evaluator's group returns time series.
    Tests:
    - Spearman correlation between group index and group mean returns (per date), then t-test over time.
    - Per-date OLS slope of group returns on group index; report mean slope and NW t-stat.
    - Jonckheere–Terpstra-like proxy using Kendall tau between group index and returns (per date).
      (True JT test requires raw observations; here we approximate using ordinal association.)
    """
    gr = e.group_returns.get(factor_name)
    if gr is None or gr.empty:
        return {
            "mono_spearman_mean": np.nan,
            "mono_spearman_t": np.nan,
            "mono_spearman_win_rate": np.nan,
            "mono_slope_mean": np.nan,
            "mono_slope_t_nw": np.nan,
            "mono_kendall_mean": np.nan,
            "mono_kendall_t": np.nan,
        }

    # Convert to date x group matrix
    gr_mat = gr.groupby(level=0).mean()
    cols = _extract_factor_group_columns(gr_mat, factor_name, n_groups)
    gr_mat = gr_mat[cols].copy()

    if use_excess:
        gr_mat = gr_mat.sub(gr_mat.mean(axis=1), axis=0)

    x = np.arange(1, n_groups + 1)

    spears = []
    slopes = []
    kendalls = []

    for dt, row in gr_mat.iterrows():
        y = row.values.astype(float)
        if np.all(np.isnan(y)):
            continue
        if np.sum(np.isfinite(y)) < max(5, n_groups // 2):
            continue

        r_s, _ = stats.spearmanr(x, y, nan_policy="omit")
        spears.append(r_s)

        # OLS slope of y on x
        mask = np.isfinite(y)
        if mask.sum() >= 3:
            X = sm.add_constant(x[mask])
            fit = sm.OLS(y[mask], X).fit()
            slopes.append(fit.params[1])
        else:
            slopes.append(np.nan)

        r_k, _ = stats.kendalltau(x, y, nan_policy="omit")
        kendalls.append(r_k)

    spears = pd.Series(spears).dropna()
    slopes = pd.Series(slopes).dropna()
    kendalls = pd.Series(kendalls).dropna()

    def mean_t(s: pd.Series) -> Tuple[float, float]:
        if len(s) < 30:
            return (safe_float(s.mean()), np.nan)
        mu = float(s.mean())
        sd = float(s.std(ddof=1))
        t = mu / (sd / np.sqrt(len(s))) if sd > 0 else np.nan
        return mu, t

    spearman_mean, spearman_t = mean_t(spears)
    kendall_mean, kendall_t = mean_t(kendalls)
    slope_mean = safe_float(slopes.mean())
    slope_t_nw = newey_west_tstat(slopes)

    return {
        "mono_spearman_mean": spearman_mean,
        "mono_spearman_t": spearman_t,
        "mono_spearman_win_rate": (
            safe_float((spears > 0).mean()) if len(spears) else np.nan
        ),
        "mono_slope_mean": slope_mean,
        "mono_slope_t_nw": slope_t_nw,
        "mono_kendall_mean": kendall_mean,
        "mono_kendall_t": kendall_t,
    }


def run_evaluator(evaluator: Evaluator, backtest_params: BacktestParams):
    factor_names = evaluator._factor_df.columns.to_list()
    result: Dict[str, Any] = {}

    # Coverage
    evaluator.get_coverage()
    result["mean_coverage"] = evaluator.factor_coverage.mean()

    # IC and IC tests
    evaluator.get_info_coef(
        method=backtest_params.ic_method,
    )
    ic_df = evaluator.info_coef.copy()
    result["ic_raw"] = ic_df

    ic_summ_rows = []
    ic_stab_rows = []
    for col in ic_df.columns:
        s = ic_df[col]
        summ = ic_summary_stats(s)
        stab = ic_stability_tests(
            s,
            backtest_params.ic_roll_window,
            backtest_params.ic_acf_lags,
            backtest_params.ic_break_k,
        )
        ic_summ_rows.append(pd.Series(summ, name=col))
        ic_stab_rows.append(pd.Series(stab, name=col))

    result["ic_summary"] = pd.DataFrame(ic_summ_rows)
    result["ic_stability"] = pd.DataFrame(ic_stab_rows)

    # Group returns
    evaluator.get_group_returns(
        n=backtest_params.n_groups,
        mode=backtest_params.bucketing_mode,
    )

    factor_return = evaluator.sorted_factor_return
    group_returns = pd.concat(
        [gr.groupby(level=0).mean() for gr in evaluator.group_returns.values()]
        + [factor_return],
        axis=1,
    )
    group_values = (group_returns.fillna(0) + 1).cumprod()
    group_values.index.name = "date"
    group_performance = group_values.apply(quool.Evaluator.evaluate)

    group_returns_mean = group_returns.mean()
    group_returns_t = group_returns_mean / (
        group_returns.std(ddof=1) / np.sqrt(group_returns.shape[0])
    )
    result["group_return_summary"] = pd.concat(
        [group_returns_mean, group_returns_t],
        axis=1,
        keys=["mean", "t"],
    )
    result["group_values"] = group_values
    result["group_performance"] = group_performance

    # Monotonicity tests (per factor)
    mono_rows = []
    for factor_name in factor_names:
        mono = group_monotonicity_tests(
            evaluator,
            factor_name=factor_name,
            n_groups=backtest_params.n_groups,
            use_excess=backtest_params.monotonicity_use_excess,
        )
        mono_rows.append(pd.Series(mono, name=factor_name))
    result["monotonicity"] = pd.DataFrame(mono_rows)
    result["factor"] = evaluator._factor_df
    return result


def run_test(
    factor_source: DuckPQSource,
    data_source: DuckPQSource,
    backtest_params: BacktestParams,
    logger: Logger,
):
    factor_data = factor_source.load(
        backtest_params.factor_paths,
        begin=backtest_params.begin,
        end=backtest_params.end,
    )
    logger.info(f"Loaded factor_data {factor_data.shape}")

    target_price = factor_source.load(
        backtest_params.target_path,
        begin=backtest_params.begin,
        end=backtest_params.end,
        pad_end=backtest_params.horizon + 1,
    ).iloc[:, 0]

    weight = None
    if backtest_params.weight_path is not None:
        weight = factor_source.load(
            backtest_params.weight_path,
            begin=backtest_params.begin,
            end=backtest_params.end,
        )
    feasible = get_feasible(
        data_source=data_source,
        begin=backtest_params.begin,
        end=target_price.index.levels[0].max(),
        min_list_days=backtest_params.min_list_days,
    ).iloc[:, 0]
    logger.info(f"Feasible and weighted loaded")

    target_price = target_price.where(feasible)
    future_return = (
        target_price.groupby("code").shift(-1 - backtest_params.horizon)
        / target_price.groupby("code").shift(-1)
        - 1
    )
    future_return = future_return.loc[
        future_return.index.levels[0][:: backtest_params.horizon], :
    ].loc[backtest_params.begin : backtest_params.end]
    logger.info(f"Loaded future_return {future_return.shape}")

    evaluator = Evaluator(
        factor=factor_data,
        future=future_return,
        weight=weight,
        feasible=feasible,
        logger=logger,
    )

    result: Dict[str, Dict[str, Any]] = {}
    result["raw"] = run_evaluator(evaluator, backtest_params)

    # Incremental tests vs baseline factors
    if backtest_params.baseline_factors:
        # Build baseline evaluator (regression with baseline factors only)
        baseline_factor_data = factor_source.load(
            backtest_params.baseline_factors,
            begin=backtest_params.begin,
            end=backtest_params.end,
        )
        residual = []
        for factor_name in factor_data.columns:
            logger.info(f"Orthogonalization for {factor_name}")
            evaluator_base = Evaluator(
                factor=baseline_factor_data,
                future=factor_data,
            )

            # Incremental cross sectional regression
            evaluator_base.cross_sectional_regression(
                add_intercept=backtest_params.cs_add_intercept,
                cov_type=backtest_params.cs_cov_type,
                white_type=backtest_params.cs_white_type,
            )
            residual.append(
                factor_data.sub(
                    (
                        baseline_factor_data
                        * evaluator_base.factor_premia[baseline_factor_data.columns]
                    )
                    .sum(axis=1)
                    .add(evaluator_base.factor_premia["intercept"]),
                    axis=0,
                )
            )

        residual = pd.concat(residual, axis=1).add_suffix("_ortho")
        evaluator_residual = Evaluator(
            factor=residual,
            future=future_return,
            feasible=feasible,
            weight=weight,
            logger=logger,
        )
        result["baseline_factor"] = baseline_factor_data
        result["inc"] = run_evaluator(
            evaluator_residual, backtest_params=backtest_params
        )

    else:
        result["baseline_factor"] = None
        result["inc"] = None

    return result


def run(
    factor_source: DuckPQSource,
    data_source: DuckPQSource,
    backtest_params: List[BacktestParams],
) -> Dict[str, Any]:
    logger = setup_logger("FactorEvaluator")
    results: Dict[str, Dict[str, Any]] = {}
    for i, params in enumerate(backtest_params):
        key = grenerate_test_key(params)
        logger.info(f"Evaluator start <{key}>")
        results[key] = run_test(
            factor_source=factor_source,
            data_source=data_source,
            backtest_params=params,
            logger=logger,
        )
        if i > 0:
            del results[key]["baseline_factor"]
    factor_total = pd.concat(
        [list(results.values())[0]["baseline_factor"]]
        + [res["raw"]["factor"] for res in results.values()]
        + (
            [res["inc"]["factor"] for res in results.values() if res["inc"] is not None]
        ),
        axis=1,
    )
    results["correlation"] = factor_total.groupby("date").corr().groupby(level=1).mean()
    results["correlation"] = results["correlation"].reindex(
        results["correlation"].columns
    )
    return results


def _to_1col_df(x, col_name: str) -> pd.DataFrame:
    """Series/array-like -> DataFrame(1 col). If already DF, return as-is."""
    if x is None:
        return pd.DataFrame()
    if isinstance(x, pd.DataFrame):
        return x.copy()
    if isinstance(x, pd.Series):
        return x.to_frame(col_name)
    # scalar
    if np.isscalar(x):
        return pd.DataFrame({col_name: [x]})
    # list/np array
    return pd.DataFrame({col_name: list(x)})


def save(save_path: Union[Path, str], results: Dict[str, Dict]):
    blocks: List[pd.DataFrame] = []

    for i, (test_key, res) in enumerate(results.items(), start=1):
        if test_key == "correlation":
            continue
        for restype in ["raw", "inc"]:
            rr = res[restype]
            if res[restype] is None:
                continue
            cov = rr.get("mean_coverage")  # often Series
            ic_summ = rr.get("ic_summary")  # DF
            ic_stab = rr.get("ic_stability")  # DF
            mono = rr.get("monotonicity")  # DF
            group_ret = rr.get("group_return_summary")  # DF
            cov_df = _to_1col_df(cov, "coverage")
            ic_summ_df = (
                ic_summ.copy() if isinstance(ic_summ, pd.DataFrame) else pd.DataFrame()
            )
            ic_stab_df = (
                ic_stab.copy() if isinstance(ic_stab, pd.DataFrame) else pd.DataFrame()
            )
            mono_df = mono.copy() if isinstance(mono, pd.DataFrame) else pd.DataFrame()
            group_ret.index = [
                f"factor_return({i})" for i in range(1, len(group_ret))
            ] + ["factor_return"]
            gr_df = group_ret.loc[:, "mean"].add_suffix("_mean")
            base_index = None
            for df_ in [ic_summ_df, ic_stab_df, mono_df, cov_df, group_ret]:
                if isinstance(df_, pd.DataFrame) and len(df_) > 0:
                    base_index = df_.index
                    break
            if base_index is None:
                continue
            gr_df = _to_1col_df(gr_df, base_index).T
            big = pd.concat(
                [cov_df, ic_summ_df, ic_stab_df, mono_df, gr_df],
                axis=1,
                join="outer",
            )
            big.insert(0, "test_key", test_key)
            big = big.reset_index().rename(columns={"index": "factor"})
            big.insert(2, "test_index", i)
            blocks.append(big)

    if not blocks:
        return pd.DataFrame()

    summary = pd.concat(blocks, axis=0, ignore_index=True)
    summary.index = pd.RangeIndex(1, len(summary) + 1, step=1, name="index")
    correlation = results["correlation"]

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        # Drawing charts in original excel file
        wb = writer.book

        summary.to_excel(writer, sheet_name="summary", index=False)

        # Add average group return bar chart
        ws = wb[f"summary"]
        bar = BarChart()
        bar.width = 30
        bar.height = 20
        bar.type = "col"
        bar.title = f"Average Return For Each Group"
        bar.y_axis.title = "Return"
        bar.x_axis.title = "Group"

        raw_data = Reference(
            ws,
            min_col=32,
            max_col=ws.max_column - 1,
            min_row=1,
            max_row=ws.max_row,
        )
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        bar.add_data(raw_data, titles_from_data=True)
        bar.set_categories(cats)
        bar.gapWidth = 50
        ws.add_chart(bar, f"A{len(summary) + 3}")

        correlation.to_excel(writer, sheet_name="correlation")
        # Correlation rule color
        ws = wb[f"correlation"]
        min_row, min_col = 2, 2
        max_row, max_col = ws.max_row, ws.max_column
        rule = ColorScaleRule(
            start_type="num",
            start_value=-1,
            start_color="2F5597",
            mid_type="num",
            mid_value=0,
            mid_color="FFFFFF",
            end_type="num",
            end_value=1,
            end_color="C00000",
        )
        cell_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        ws.conditional_formatting.add(cell_range, rule)
        ws.freeze_panes = "B2"

        for i, (test_key, res) in enumerate(results.items(), start=1):
            if test_key == "correlation":
                continue

            rr = res["raw"]
            ri = res["inc"]

            # IC data
            ic_raw = pd.concat(
                [rr["ic_raw"]] + ([] if ri is None else [ri["ic_raw"]]), axis=1
            )
            ic_raw = pd.concat([ic_raw, ic_raw.cumsum().add_suffix("_cumsum")], axis=1)
            ic_raw.to_excel(writer, sheet_name=f"info_coef_{i}")
            # IC bar chart + line chart
            factor_cols = ic_raw.columns[~ic_raw.columns.str.endswith("_cumsum")]
            ws = wb[f"info_coef_{i}"]
            for ii, col in enumerate(factor_cols):
                bar = BarChart()
                bar.width = 30
                bar.height = 20
                bar.type = "col"
                bar.title = f"Information Coefficient {col}"
                bar.y_axis.title = "IC"
                bar.x_axis.title = "Date"

                raw_data = Reference(
                    ws,
                    min_col=2 + ii,
                    max_col=2 + ii,
                    min_row=1,
                    max_row=ws.max_row,
                )
                cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                bar.add_data(raw_data, titles_from_data=True)
                bar.set_categories(cats)
                bar.gapWidth = 50

                line = LineChart()
                cum_data = Reference(
                    ws,
                    min_col=2 + len(factor_cols) + ii,
                    max_col=2 + len(factor_cols) + ii,
                    min_row=1,
                    max_row=ws.max_row,
                )
                line.add_data(cum_data, titles_from_data=True)
                for s in line.series:
                    s.graphicalProperties.line = LineProperties(w=int(1 * 12700))
                line.set_categories(cats)

                line.y_axis.axId = 200
                line.y_axis.crosses = "max"
                line.y_axis.title = "Cumulative IC"

                bar += line
                ws.add_chart(bar, f"B{2 + 50*ii}")

            # Group Data
            group_values = pd.concat(
                [rr["group_values"]] + ([] if ri is None else [ri["group_values"]]),
                axis=1,
            )
            group_values.to_excel(writer, sheet_name=f"group_value_{i}")
            # Group value line chart
            ws = wb[f"group_value_{i}"]
            max_row = ws.max_row
            max_col = ws.max_column
            n = group_values.shape[1] // len(factor_cols) - 1
            for ii, fcol in enumerate(factor_cols):
                chart = LineChart()
                chart.width = 30
                chart.height = 20
                chart.title = f"{fcol} Group Net Value"
                chart.y_axis.title = "Net Value"

                cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)

                data = Reference(
                    ws,
                    min_col=ii * n + 2 + ii,
                    min_row=1,
                    max_col=(ii + 1) * n + 1,
                    max_row=max_row,
                )

                chart.add_data(data, titles_from_data=True)
                for s in chart.series:
                    s.graphicalProperties.line = LineProperties(w=int(1 * 12700))
                chart.set_categories(cats)

                chart_right = LineChart()
                data = Reference(
                    ws,
                    min_col=(ii + 1) * n + 2 + ii,
                    min_row=1,
                    max_col=(ii + 1) * n + 2 + ii,
                    max_row=max_row,
                )
                chart_right.add_data(data, titles_from_data=True)
                chart_right.set_categories(cats)
                for s in chart_right.series:
                    s.graphicalProperties.line = LineProperties(w=int(1 * 12700))

                chart_right.y_axis.axId = 200
                chart_right.y_axis.crosses = "max"
                chart_right.y_axis.title = "Cumulative IC"

                chart += chart_right
                ws.add_chart(chart, f"B{2 + 50*ii}")

            # Group performance
            group_performance = pd.concat(
                [rr["group_performance"]]
                + ([] if ri is None else [ri["group_performance"]]),
                axis=1,
            )
            group_performance.to_excel(writer, sheet_name=f"group_performance_{i}")


if __name__ == "__main__":
    DATASET_PATH = os.getenv("DATASET_PATH")
    FACTOR_DATA_PATH = os.getenv("FACTOR_DATA_PATH")
    if not DATASET_PATH or not FACTOR_DATA_PATH:
        raise EnvironmentError(
            "Missing env vars: DATASET_PATH and/or FACTOR_DATA_PATH. "
            "Please set them (e.g. in .env) before running."
        )
    data_source = DuckPQSource(Path(DATASET_PATH))
    data_source.register("quotes_day")
    data_source.register("instruments_info")

    params = [
        BacktestParams(
            factor_paths=f"example_factor/{name}",  # CONFIG: Placeholder for factor paths
            baseline_factors=[
                "barra/size",
                "barra/momentum",
                "barra/volatility",
                "barra/liquidity",
                "barra/value",
                "barra/profitability",
                "barra/leverage",
                "barra/growth",
            ],
            begin="2015-01-01",
            end="2025-06-30",
        )
        for name in ["example_factor_name"]
    ]

    output_path = "highfreq_skew_kurt__batch2.xlsx"  # CONFIG: Placeholder for output path
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    factor_source = DuckPQSource(Path(FACTOR_DATA_PATH))
    results = run(factor_source, data_source, params)
    save(output_path, results)
