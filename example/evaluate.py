import os
import dotenv
from pathlib import Path
from logging import Logger
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple, Union, Optional

import numpy as np
import pandas as pd
import statsmodels.api as sm
from scipy import stats
from statsmodels.stats.diagnostic import acorr_ljungbox

import quool
from parquool import setup_logger

from factool import DuckPQSource, Evaluator


dotenv.load_dotenv()


FactorPath = str


def _safe_float(x: Any) -> float:
    try:
        if x is None:
            return np.nan
        if isinstance(x, (np.floating, float, int, np.integer)):
            return float(x)
        return float(np.asarray(x))
    except Exception:
        return np.nan


def _get_feasible(source: DuckPQSource, begin: str, end: str, min_list_days: int = 90):
    source.register("instruments_info")
    sql = f"""
    SELECT
        q.date AS date,
        q.code AS code,
        (
            q.high > q.limit_down
            AND q.low  < q.limit_up
            AND COALESCE(q.st, false) = false
            AND COALESCE(q.suspended, false) = false
            AND datediff('day', i.listed_date, q.date) > {min_list_days}
        ) AS feasible
    FROM quotes_day AS q
    JOIN instruments_info AS i
        ON q.code = i.code
    WHERE q.date >= '{begin}' AND q.date <= '{end}'
    """

    data = source.query(sql)
    data["date"] = pd.to_datetime(data["date"])

    feasible = data.set_index(["date", "code"]).sort_index()
    return feasible


@dataclass
class BacktestParams:
    factor_paths: List[str]
    begin: str
    end: str
    target_path: str = "quotes_day/open_post"
    horizon: int = 5
    baseline_factors: Optional[Union[str, List[str]]] = None

    # Tradability mask and weight
    min_list_days: int = 90
    weight_path: str = None

    # IC and grouping
    ic_method: str = "spearman"
    n_groups: int = 10
    bucketing_mode: str = "single"

    # Cross-sectional regression
    cs_add_intercept: bool = True
    cs_cov_type: str = "white"
    cs_white_type: str = "HC1"

    # IC stability settings
    ic_roll_window: int = 60
    ic_acf_lags: int = 10
    ic_break_k: int = (
        1  # number of candidate breakpoints to test (simple single-break scan)
    )

    # Monotonicity settings
    monotonicity_use_excess: bool = (
        False  # if True, use group returns minus cross-sectional mean
    )


def _generate_test_key(param: BacktestParams) -> str:
    paths = (
        param.factor_paths
        if isinstance(param.factor_paths, list)
        else [param.factor_paths]
    )
    bases = []
    if param.baseline_factors is not None:
        bases = (
            param.baseline_factors
            if isinstance(param.baseline_factors, list)
            else [param.baseline_factors]
        )
    return "+".join([path.split("/", 1)[-1] for path in paths]) + (
        ""
        if not bases
        else (f"__" + "+".join([base.split("/", 1)[-1] for base in bases]))
    )


def _newey_west_tstat(x: pd.Series, lags: Optional[int] = None) -> float:
    """
    Newey-West t-stat for mean(x) with HAC covariance.
    Uses statsmodels OLS of x on constant.
    """
    s = pd.Series(x).dropna()
    if s.shape[0] < 10:
        return np.nan
    y = s.values
    X = np.ones((len(y), 1))
    model = sm.OLS(y, X)
    if lags is None:
        # Common rule-of-thumb
        lags = int(np.floor(4 * (len(y) / 100) ** (2 / 9)))
    try:
        fit = model.fit(cov_type="HAC", cov_kwds={"maxlags": int(lags)})
        return float(fit.tvalues[0])
    except Exception:
        return np.nan


def _ic_summary_stats(ic: pd.Series) -> Dict[str, Any]:
    """
    IC summary including win rate, ICIR, tails, and HAC t-stat.
    """
    s = pd.Series(ic).dropna()
    n = int(s.shape[0])
    if n == 0:
        return {
            "ic_n": 0,
            "ic_mean": np.nan,
            "ic_std": np.nan,
            "ic_ir": np.nan,
            "ic_t": np.nan,
            "ic_t_nw": np.nan,
            "ic_win_rate": np.nan,
            "ic_p05": np.nan,
            "ic_p50": np.nan,
            "ic_p95": np.nan,
            "ic_skew": np.nan,
            "ic_kurt": np.nan,
            "ic_tail_gt_2std": np.nan,
        }

    mu = float(s.mean())
    sd = float(s.std(ddof=1)) if n > 1 else np.nan
    ir = mu / sd if sd and np.isfinite(sd) and sd > 0 else np.nan
    t = mu / (sd / np.sqrt(n)) if sd and np.isfinite(sd) and sd > 0 else np.nan
    t_nw = _newey_west_tstat(s)

    win_rate = float((s > 0).mean())
    p05, p50, p95 = [float(v) for v in np.nanpercentile(s.values, [5, 50, 95])]
    skew = float(stats.skew(s.values, nan_policy="omit")) if n >= 3 else np.nan
    kurt = (
        float(stats.kurtosis(s.values, fisher=True, nan_policy="omit"))
        if n >= 4
        else np.nan
    )
    tail_gt_2std = float((np.abs(s - mu) > 2 * sd).mean()) if sd and sd > 0 else np.nan

    return {
        "ic_n": n,
        "ic_mean": mu,
        "ic_std": sd,
        "ic_ir": ir,
        "ic_t": t,
        "ic_t_nw": t_nw,
        "ic_win_rate": win_rate,
        "ic_p05": p05,
        "ic_p50": p50,
        "ic_p95": p95,
        "ic_skew": skew,
        "ic_kurt": kurt,
        "ic_tail_gt_2std": tail_gt_2std,
    }


def _ic_stability_tests(
    ic: pd.Series, roll_window: int, acf_lags: int, break_k: int = 1
) -> Dict[str, Any]:
    """
    Stability/drift tests:
    - Rolling mean/IR variability
    - ACF(1) and Ljung-Box p-value (autocorrelation)
    - Simple single-break scan (max abs difference in pre/post mean), report best split and p-value
    """
    s = pd.Series(ic).dropna()
    out: Dict[str, Any] = {}
    if s.shape[0] < max(50, roll_window + 5):
        out.update(
            {
                "ic_roll_mean_std": np.nan,
                "ic_roll_ir_std": np.nan,
                "ic_acf1": np.nan,
                "ic_lb_pvalue": np.nan,
                "ic_best_break_date": None,
                "ic_break_stat": np.nan,
                "ic_break_pvalue": np.nan,
            }
        )
        return out

    roll_mean = s.rolling(roll_window).mean()
    roll_std = s.rolling(roll_window).std(ddof=1)
    roll_ir = roll_mean / roll_std

    out["ic_roll_mean_std"] = _safe_float(roll_mean.dropna().std(ddof=1))
    out["ic_roll_ir_std"] = _safe_float(roll_ir.dropna().std(ddof=1))

    # Autocorrelation tests
    try:
        out["ic_acf1"] = _safe_float(s.autocorr(lag=1))
    except Exception:
        out["ic_acf1"] = np.nan

    try:
        lb = acorr_ljungbox(s.values, lags=[min(acf_lags, len(s) - 1)], return_df=True)
        out["ic_lb_pvalue"] = _safe_float(lb["lb_pvalue"].iloc[-1])
    except Exception:
        out["ic_lb_pvalue"] = np.nan

    # Simple best single break scan (choose split that maximizes |mean1 - mean2|)
    idx = s.index
    values = s.values
    n = len(s)
    min_seg = max(30, n // 10)
    best = {"stat": -np.inf, "split": None, "pvalue": np.nan}

    # Candidate splits
    split_candidates = list(range(min_seg, n - min_seg))
    if break_k is not None and break_k > 0 and len(split_candidates) > 0:
        # To limit compute, subsample candidates
        step = max(1, len(split_candidates) // 200)
        split_candidates = split_candidates[::step]

    for k in split_candidates:
        a = values[:k]
        b = values[k:]
        if len(a) < min_seg or len(b) < min_seg:
            continue
        # Two-sample t-test (Welch)
        stat, pval = stats.ttest_ind(a, b, equal_var=False, nan_policy="omit")
        diff = np.nanmean(a) - np.nanmean(b)
        score = np.abs(diff)
        if np.isfinite(score) and score > best["stat"]:
            best = {"stat": score, "split": k, "pvalue": pval}

    if best["split"] is None:
        out["ic_best_break_date"] = None
        out["ic_break_stat"] = np.nan
        out["ic_break_pvalue"] = np.nan
    else:
        out["ic_best_break_date"] = str(idx[best["split"]].date())
        out["ic_break_stat"] = _safe_float(best["stat"])
        out["ic_break_pvalue"] = _safe_float(best["pvalue"])

    return out


def _extract_factor_group_columns(
    group_returns: pd.DataFrame, factor_name: str, n_groups: int
) -> List[str]:
    return [f"{factor_name}({i + 1})" for i in range(n_groups)]


def _group_monotonicity_tests(
    e: Evaluator,
    factor_name: str,
    n_groups: int,
    use_excess: bool = False,
) -> Dict[str, Any]:
    """
    Uses evaluator's group returns time series.
    Tests:
    - Spearman correlation between group index and group mean returns (per date), then t-test over time.
    - Per-date OLS slope of group returns on group index; report mean slope and NW t-stat.
    - Jonckheere–Terpstra-like proxy using Kendall tau between group index and returns (per date).
      (True JT test requires raw observations; here we approximate using ordinal association.)
    """
    gr = e.group_returns.get(factor_name)
    if gr is None or gr.empty:
        return {
            "mono_spearman_mean": np.nan,
            "mono_spearman_t": np.nan,
            "mono_spearman_win_rate": np.nan,
            "mono_slope_mean": np.nan,
            "mono_slope_t_nw": np.nan,
            "mono_kendall_mean": np.nan,
            "mono_kendall_t": np.nan,
        }

    # Convert to date x group matrix
    gr_mat = gr.groupby(level=0).mean()
    cols = _extract_factor_group_columns(gr_mat, factor_name, n_groups)
    gr_mat = gr_mat[cols].copy()

    if use_excess:
        gr_mat = gr_mat.sub(gr_mat.mean(axis=1), axis=0)

    x = np.arange(1, n_groups + 1)

    spears = []
    slopes = []
    kendalls = []

    for dt, row in gr_mat.iterrows():
        y = row.values.astype(float)
        if np.all(np.isnan(y)):
            continue
        if np.sum(np.isfinite(y)) < max(5, n_groups // 2):
            continue

        r_s, _ = stats.spearmanr(x, y, nan_policy="omit")
        spears.append(r_s)

        # OLS slope of y on x
        mask = np.isfinite(y)
        if mask.sum() >= 3:
            X = sm.add_constant(x[mask])
            fit = sm.OLS(y[mask], X).fit()
            slopes.append(fit.params[1])
        else:
            slopes.append(np.nan)

        r_k, _ = stats.kendalltau(x, y, nan_policy="omit")
        kendalls.append(r_k)

    spears = pd.Series(spears).dropna()
    slopes = pd.Series(slopes).dropna()
    kendalls = pd.Series(kendalls).dropna()

    def mean_t(s: pd.Series) -> Tuple[float, float]:
        if len(s) < 30:
            return (_safe_float(s.mean()), np.nan)
        mu = float(s.mean())
        sd = float(s.std(ddof=1))
        t = mu / (sd / np.sqrt(len(s))) if sd > 0 else np.nan
        return mu, t

    spearman_mean, spearman_t = mean_t(spears)
    kendall_mean, kendall_t = mean_t(kendalls)
    slope_mean = _safe_float(slopes.mean())
    slope_t_nw = _newey_west_tstat(slopes)

    return {
        "mono_spearman_mean": spearman_mean,
        "mono_spearman_t": spearman_t,
        "mono_spearman_win_rate": (
            _safe_float((spears > 0).mean()) if len(spears) else np.nan
        ),
        "mono_slope_mean": slope_mean,
        "mono_slope_t_nw": slope_t_nw,
        "mono_kendall_mean": kendall_mean,
        "mono_kendall_t": kendall_t,
    }


def _run_evaluator(evaluator: Evaluator, backtest_params: BacktestParams):
    factor_names = evaluator._factor_df.columns.to_list()
    result: Dict[str, Any] = {}

    # Coverage
    evaluator.get_coverage()
    result["mean_coverage"] = evaluator.factor_coverage.mean()

    # IC and IC tests
    evaluator.get_info_coef(method=backtest_params.ic_method)
    ic_df = evaluator.info_coef.copy()
    result["ic_raw"] = ic_df

    ic_summ_rows, ic_stab_rows = [], []
    for col in ic_df.columns:
        s = ic_df[col]
        summ = _ic_summary_stats(s)
        stab = _ic_stability_tests(
            s,
            backtest_params.ic_roll_window,
            backtest_params.ic_acf_lags,
            backtest_params.ic_break_k,
        )
        ic_summ_rows.append(pd.Series(summ, name=col))
        ic_stab_rows.append(pd.Series(stab, name=col))

    result["ic_summary"] = pd.DataFrame(ic_summ_rows)
    result["ic_stability"] = pd.DataFrame(ic_stab_rows)

    # Group returns
    evaluator.get_group_returns(
        n=backtest_params.n_groups, mode=backtest_params.bucketing_mode
    )

    factor_return = evaluator.sorted_factor_return
    group_returns = pd.concat(
        [gr.groupby(level=0).mean() for gr in evaluator.group_returns.values()]
        + [factor_return],
        axis=1,
    )
    group_returns.index.name = "date"

    result["group_returns"] = group_returns

    group_values = (group_returns.fillna(0) + 1).cumprod()
    group_values.index.name = "date"

    group_performance = group_values.apply(quool.Evaluator.evaluate)
    group_returns_mean = group_returns.mean()
    group_returns_t = group_returns_mean / (
        group_returns.std(ddof=1) / np.sqrt(group_returns.shape[0])
    )
    result["group_return_summary"] = pd.concat(
        [group_returns_mean, group_returns_t],
        axis=1,
        keys=["mean", "t"],
    )
    result["group_values"] = group_values
    result["group_performance"] = group_performance

    # Monotonicity tests (per factor)
    mono_rows = []
    for factor_name in factor_names:
        mono = _group_monotonicity_tests(
            evaluator,
            factor_name=factor_name,
            n_groups=backtest_params.n_groups,
            use_excess=backtest_params.monotonicity_use_excess,
        )
        mono_rows.append(pd.Series(mono, name=factor_name))
    result["monotonicity"] = pd.DataFrame(mono_rows)

    result["factor"] = evaluator._factor_df
    return result


def _run_test(
    factor_source: DuckPQSource,
    backtest_params: BacktestParams,
    logger: Logger,
):
    factor_data = factor_source.load(
        backtest_params.factor_paths,
        begin=backtest_params.begin,
        end=backtest_params.end,
    )
    logger.info(f"Loaded factor_data {factor_data.shape}")

    target_price = factor_source.load(
        backtest_params.target_path,
        begin=backtest_params.begin,
        end=backtest_params.end,
        pad_end=backtest_params.horizon + 1,
    ).iloc[:, 0]

    weight = None
    if backtest_params.weight_path is not None:
        weight = factor_source.load(
            backtest_params.weight_path,
            begin=backtest_params.begin,
            end=backtest_params.end,
        )
    feasible = _get_feasible(
        source=factor_source,
        begin=backtest_params.begin,
        end=target_price.index.levels[0].max(),
        min_list_days=backtest_params.min_list_days,
    ).iloc[:, 0]
    logger.info(f"Feasible and weighted loaded")

    target_price = target_price.where(feasible)
    future_return = (
        target_price.groupby("code").shift(-1 - backtest_params.horizon)
        / target_price.groupby("code").shift(-1)
        - 1
    )
    future_return = future_return.loc[
        future_return.index.levels[0][:: backtest_params.horizon], :
    ].loc[backtest_params.begin : backtest_params.end]
    logger.info(f"Loaded future_return {future_return.shape}")

    evaluator = Evaluator(
        factor=factor_data,
        future=future_return,
        weight=weight,
        feasible=feasible,
        logger=logger,
    )

    result: Dict[str, Dict[str, Any]] = {}
    result["raw"] = _run_evaluator(evaluator, backtest_params)

    # Incremental tests vs baseline factors
    if backtest_params.baseline_factors:
        # Build baseline evaluator (regression with baseline factors only)
        baseline_factor_data = factor_source.load(
            backtest_params.baseline_factors,
            begin=backtest_params.begin,
            end=backtest_params.end,
        )
        residual = []
        for factor_name in factor_data.columns:
            logger.info(f"Orthogonalization for {factor_name}")
            evaluator_base = Evaluator(
                factor=baseline_factor_data,
                future=factor_data,
            )

            # Incremental cross sectional regression
            evaluator_base.cross_sectional_regression(
                add_intercept=backtest_params.cs_add_intercept,
                cov_type=backtest_params.cs_cov_type,
                white_type=backtest_params.cs_white_type,
            )
            residual.append(
                factor_data.sub(
                    (
                        baseline_factor_data
                        * evaluator_base.factor_premia[baseline_factor_data.columns]
                    )
                    .sum(axis=1)
                    .add(evaluator_base.factor_premia["intercept"]),
                    axis=0,
                )
            )

        residual = pd.concat(residual, axis=1).add_suffix("_ortho")
        evaluator_residual = Evaluator(
            factor=residual,
            future=future_return,
            feasible=feasible,
            weight=weight,
            logger=logger,
        )
        result["baseline_factor"] = baseline_factor_data
        result["inc"] = _run_evaluator(
            evaluator_residual, backtest_params=backtest_params
        )

    else:
        result["baseline_factor"] = None
        result["inc"] = None

    return result


def run(
    factor_source: DuckPQSource,
    backtest_params: List[BacktestParams],
) -> Dict[str, Any]:
    logger = setup_logger("FactorEvaluator")
    results: Dict[str, Dict[str, Any]] = {}

    for i, params in enumerate(backtest_params):
        key = _generate_test_key(params)
        logger.info(f"Evaluator start <{key}>")
        results[key] = _run_test(
            factor_source=factor_source,
            backtest_params=params,
            logger=logger,
        )
        if i > 0:
            del results[key]["baseline_factor"]

    factor_total = pd.concat(
        [list(results.values())[0]["baseline_factor"]]
        + [res["raw"]["factor"] for res in results.values()]
        + (
            [res["inc"]["factor"] for res in results.values() if res["inc"] is not None]
        ),
        axis=1,
    )

    # index: (date, factor1), columns: factor2
    corr_ts = factor_total.groupby("date").corr()
    corr_ts = corr_ts.reindex(corr_ts.columns, level=1)
    results["correlation"] = corr_ts

    return results


def _to_1col_df(x, col_name: str) -> pd.DataFrame:
    """Series/array-like -> DataFrame(1 col). If already DF, return as-is."""
    if x is None:
        return pd.DataFrame()
    if isinstance(x, pd.DataFrame):
        return x.copy()
    if isinstance(x, pd.Series):
        return x.to_frame(col_name)
    # scalar
    if np.isscalar(x):
        return pd.DataFrame({col_name: [x]})
    # list/np array
    return pd.DataFrame({col_name: list(x)})


def _clean(results: Dict[str, Dict]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    blocks: List[pd.DataFrame] = []

    for i, (test_key, res) in enumerate(results.items(), start=1):
        if test_key == "correlation":
            continue

        for restype in ["raw", "inc"]:
            rr = res.get(restype)
            if rr is None:
                continue

            cov = rr.get("mean_coverage")
            ic_summ = rr.get("ic_summary")
            ic_stab = rr.get("ic_stability")
            mono = rr.get("monotonicity")

            cov_df = _to_1col_df(cov, "coverage")
            ic_summ_df = (
                ic_summ.copy() if isinstance(ic_summ, pd.DataFrame) else pd.DataFrame()
            )
            ic_stab_df = (
                ic_stab.copy() if isinstance(ic_stab, pd.DataFrame) else pd.DataFrame()
            )
            mono_df = mono.copy() if isinstance(mono, pd.DataFrame) else pd.DataFrame()

            base_index = None
            for df_ in [ic_summ_df, ic_stab_df, mono_df, cov_df]:
                if isinstance(df_, pd.DataFrame) and len(df_) > 0:
                    base_index = df_.index
                    break
            if base_index is None:
                continue

            big = pd.concat(
                [cov_df, ic_summ_df, ic_stab_df, mono_df],
                axis=1,
                join="outer",
            )
            big.insert(0, "test_key", test_key)
            big = big.reset_index().rename(columns={"index": "factor"})
            big.insert(2, "test_index", i)
            big.insert(3, "restype", restype)
            blocks.append(big)

    summary = pd.concat(blocks, axis=0, ignore_index=True) if blocks else pd.DataFrame()
    if not summary.empty:
        summary.index = pd.RangeIndex(1, len(summary) + 1, step=1, name="index")

    correlation_ts = results.get("correlation", pd.DataFrame())
    return summary, correlation_ts


def show(results: Dict[str, Dict], test_key: Optional[str] = None, width: str = "100%"):
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    from IPython.display import display, Markdown
    import seaborn as sns

    def _ensure_dt_index(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or getattr(df, "empty", True):
            return df
        out = df.copy()
        if not isinstance(out.index, pd.DatetimeIndex):
            try:
                out.index = pd.to_datetime(out.index)
            except Exception:
                pass
        return out.sort_index()

    def _to_numeric_df(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or getattr(df, "empty", True):
            return df
        out = df.copy()
        for c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
        return out

    def _display_table(df: pd.DataFrame, title: str):
        display(Markdown(f"## {title}"))
        if df is None or getattr(df, "empty", True):
            display(Markdown("*(empty)*"))
            return
        display(df.style.format(precision=4, na_rep="—"))

    # notebook width
    display(
        Markdown(
            f"""
<style>
.output_wrapper, .output {{ width: {width} !important; max-width: {width} !important; }}
</style>
"""
        )
    )
    display(Markdown("# Factor Evaluation Results"))

    summary, corr_ts = _clean(results)
    _display_table(summary, "0. Summary")

    # test keys
    test_keys = [k for k in results.keys() if k not in ("correlation", "summary")]
    if test_key is not None:
        if test_key not in results:
            raise KeyError(f"test_key={test_key} not in results")
        test_keys = [test_key]

    # correlation mean matrix for heatmap
    corr_mean = None
    if isinstance(corr_ts, pd.DataFrame) and not corr_ts.empty:
        # corr_ts: index (date, factor1), columns factor2
        corr_mean = corr_ts.groupby(level=1).mean()
        corr_mean = corr_mean.reindex(corr_mean.columns)

    def _plot_ic(ic: pd.DataFrame, title: str):
        ic = _ensure_dt_index(_to_numeric_df(ic))
        if ic is None or ic.empty:
            display(Markdown("*(IC empty)*"))
            return

        n = ic.shape[1]
        fig, axes = plt.subplots(n, 1, figsize=(14, 3.2 * n), sharex=True)
        if n == 1:
            axes = [axes]

        for ax, col in zip(axes, ic.columns):
            s = ic[col].dropna()
            if s.empty:
                continue
            cs = s.cumsum()

            ax.plot(s.index, s.values, color="tab:blue", linewidth=1.0, label="IC")
            ax.axhline(0, color="gray", linestyle="--", linewidth=0.8)
            ax.set_ylabel("Daily IC")

            ax2 = ax.twinx()
            ax2.plot(
                cs.index,
                cs.values,
                color="tab:orange",
                linewidth=1.2,
                label="Cumsum IC",
            )
            ax2.set_ylabel("Cumsum IC")

            ax.set_title(f"{title} | {col}")
            # combined legend
            lines = ax.get_lines() + ax2.get_lines()
            labels = [l.get_label() for l in lines]
            ax.legend(lines, labels, loc="upper left", frameon=False)

        plt.tight_layout()
        plt.show()
    
    def _plot_ic_hist(ic: pd.DataFrame, title: str):
        ic = _ensure_dt_index(_to_numeric_df(ic))
        if ic is None or ic.empty:
            display(Markdown("*(IC empty)*"))
            return

        n = ic.shape[1]
        fig, axes = plt.subplots(n, 1, figsize=(10, 3.2 * n))
        if n == 1:
            axes = [axes]

        for ax, col in zip(axes, ic.columns):
            s = ic[col].dropna()
            if s.empty:
                continue

            sns.histplot(s.values, bins=30, kde=True, ax=ax, color="tab:blue")
            ax.set_title(f"{title} | {col}")
            ax.set_xlabel("IC Value")
            ax.set_ylabel("Frequency")

        plt.tight_layout()
        plt.show()

    def _plot_group_return_box(
        raw_gr: pd.DataFrame, inc_gr: Optional[pd.DataFrame], title: str
    ):
        # gr: date x columns (group returns)
        raw_gr = _ensure_dt_index(_to_numeric_df(raw_gr))
        inc_gr = (
            _ensure_dt_index(_to_numeric_df(inc_gr)) if inc_gr is not None else None
        )
        if raw_gr is None or raw_gr.empty:
            display(Markdown("*(group_returns empty)*"))
            return

        groups = list(raw_gr.columns)
        data = []
        positions = []
        labels = []
        width = 0.35

        for i, g in enumerate(groups):
            r = raw_gr[g].dropna().values
            data.append(r)
            positions.append(i - width / 2)
            labels.append((g, "raw"))
            if inc_gr is not None and g in inc_gr.columns:
                r2 = inc_gr[g].dropna().values
                data.append(r2)
                positions.append(i + width / 2)
                labels.append((g, "inc"))

        fig, ax = plt.subplots(figsize=(14, 5))
        bp = ax.boxplot(
            data,
            positions=positions,
            widths=0.28,
            patch_artist=True,
            showfliers=False,
        )
        
        # color raw/inc
        for patch, lab in zip(bp["boxes"], labels):
            patch.set_facecolor("tab:blue" if lab[1] == "raw" else "tab:orange")
            patch.set_alpha(0.45)

        ax.axhline(0, color="gray", linestyle="--", linewidth=0.8)
        ax.set_title(title)
        ax.set_ylabel("Group Return")
        ax.set_xticks(range(len(groups)))
        ax.set_xticklabels([str(g) for g in groups], rotation=0)
        ax.tick_params(axis="x", labelrotation=45)

        # manual legend
        handles = []
        from matplotlib.patches import Patch

        handles.append(Patch(facecolor="tab:blue", alpha=0.45, label="raw"))
        if inc_gr is not None:
            handles.append(Patch(facecolor="tab:orange", alpha=0.45, label="inc"))
        ax.legend(handles=handles, frameon=False, loc="upper left")

        plt.tight_layout()
        plt.show()

    def _plot_group_values(gv: pd.DataFrame, title: str):
        gv = _ensure_dt_index(_to_numeric_df(gv))
        if gv is None or gv.empty:
            display(Markdown("*(group_values empty)*"))
            return
        fig, ax = plt.subplots(figsize=(14, 5))
        for c in gv.columns:
            ax.plot(gv.index, gv[c], linewidth=1.0, label=str(c))
        ax.set_title(title)
        ax.set_ylabel("NAV")
        ax.legend(ncol=5, frameon=False, fontsize=9)
        plt.tight_layout()
        plt.show()

    def _plot_corr_heatmap(corr_mean: pd.DataFrame, title: str):
        if corr_mean is None or corr_mean.empty:
            display(Markdown("*(correlation empty)*"))
            return
        fig, ax = plt.subplots(figsize=(12, 10))
        mat = corr_mean.values.astype(float)
        sns.heatmap(
            corr_mean,
            ax=ax,
            cmap="RdBu_r",
            vmin=-1,
            vmax=1,
            center=0,
            square=True,
            cbar_kws={"shrink": 0.8},
        )
        ax.set_title(title)
        plt.tight_layout()
        plt.show()

    def _corr_pairs_long(corr_ts: pd.DataFrame) -> pd.DataFrame:
        # corr_ts: (date, factor1) x factor2 -> long (date, pair, corr)
        if corr_ts is None or corr_ts.empty:
            return pd.DataFrame()
        df = corr_ts.copy()
        df.index = df.index.set_names(["date", "factor1"])
        long = (
            df.stack()
            .rename("corr")
            .reset_index()
            .rename(columns={"level_2": "factor2"})
        )
        # remove diagonal and keep factor1<factor2 to avoid duplicates
        long = long[long["factor1"] != long["factor2"]].copy()
        # stable pair name
        a = long["factor1"].astype(str)
        b = long["factor2"].astype(str)
        long["pair"] = np.where(a < b, a + " ~ " + b, b + " ~ " + a)
        long = long.drop(columns=["factor1", "factor2"]).drop_duplicates(
            subset=["date", "pair"]
        )
        return long

    def _plot_corr_box(corr_ts: pd.DataFrame, title: str, max_pairs: int = 60):
        long = _corr_pairs_long(corr_ts)
        if long.empty:
            display(Markdown("*(correlation empty)*"))
            return

        # pairs too many -> keep highest |mean|
        stat = long.groupby("pair")["corr"].mean().abs().sort_values(ascending=False)
        keep = stat.head(max_pairs).index
        long = long[long["pair"].isin(keep)]

        # pivot to list of arrays for boxplot
        pairs = list(keep)
        data = [long.loc[long["pair"] == p, "corr"].dropna().values for p in pairs]

        fig, ax = plt.subplots(figsize=(14, 6))
        ax.boxplot(data, showfliers=False)
        ax.axhline(0, color="gray", linestyle="--", linewidth=0.8)
        ax.set_title(title)
        ax.set_ylabel("Correlation")
        ax.set_xticks(np.arange(1, len(pairs) + 1))
        ax.set_xticklabels(pairs, rotation=90, fontsize=8)
        plt.tight_layout()
        plt.show()

    # per test show
    for i, tk in enumerate(test_keys, start=1):
        res = results[tk]
        rr = res.get("raw") or {}
        ri = res.get("inc")

        display(Markdown(f"# {i}. {tk}"))

        # 1) IC
        display(Markdown(f"## {i}.1 IC 分析"))
        display(Markdown(f"### {i}.1.1 IC 时序图"))
        _plot_ic(rr.get("ic_raw"), title=f"{tk} | IC (raw)")
        if ri is not None:
            _plot_ic(ri.get("ic_raw"), title=f"{tk} | IC (inc)")
        display(Markdown(f"### {i}.1.2 频率分布直方图"))
        _plot_ic_hist(rr.get("ic_raw"), title=f"{tk} | IC Histogram (raw)")
        if ri is not None:
            _plot_ic_hist(ri.get("ic_raw"), title=f"{tk} | IC Histogram (inc)")

        # 2) Grouping
        display(Markdown(f"## {i}.2 分组收益分析"))
        display(Markdown(f"### {i}.2.1 分组回报箱线图"))
        raw_gr = rr.get("group_returns")
        inc_gr = None if ri is None else ri.get("group_returns")
        _plot_group_return_box(raw_gr, inc_gr, title=f"{tk} | Group Returns Boxplot")

        display(Markdown(f"### {i}.2.2 分组累积收益图（Group NAV）"))
        _plot_group_values(rr.get("group_values"), title=f"{tk} | Group NAV (raw)")
        if ri is not None:
            _plot_group_values(ri.get("group_values"), title=f"{tk} | Group NAV (inc)")

    # 3) Correlation
    display(Markdown(f"# {i + 1} 相关性分析"))
    display(Markdown(f"# {i + 1}.1 相关性热力图（时序均值）"))
    _plot_corr_heatmap(corr_mean, title="Correlation Heatmap (mean over time)")

    # 5) 相关性箱线图（使用时序相关性序列）
    display(Markdown(f"# {i + 1}.2 相关性箱线图（时序分布）"))
    _plot_corr_box(corr_ts, title="Correlation Boxplot (pairwise over time)")


def save(save_path: Union[Path, str], results: Dict[str, Dict]):
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.drawing.line import LineProperties

    summary, corr_ts = _clean(results)

    # ---- add group return mean ----
    if summary is not None and not summary.empty:
        # build mapping for quick access
        # results[test_key][restype]["group_returns"] : date x columns
        group_mean_rows = []
        for _, row in summary[["test_key", "factor", "restype"]].iterrows():
            tk = row["test_key"]
            fac = row["factor"]
            rt = row["restype"]
            rr = results[tk].get(rt)
            if rr is None:
                group_mean_rows.append({})
                continue
            gr = rr.get("group_returns")
            if gr is None or gr.empty:
                group_mean_rows.append({})
                continue
            # pick group columns for this factor (fac(1..n))
            cols = [c for c in gr.columns if str(c).startswith(f"{fac}(")]
            d = {}
            if cols:
                m = gr[cols].mean()
                for c, v in m.items():
                    d[f"{c}_mean"] = float(v)
            group_mean_rows.append(d)

        group_mean_df = pd.DataFrame(group_mean_rows)
        summary_out = pd.concat([summary.reset_index(drop=True), group_mean_df], axis=1)
    else:
        summary_out = summary

    # ---- correlation mean matrix for conditional format & heatmap-like table ----
    corr_mean = None
    if isinstance(corr_ts, pd.DataFrame) and not corr_ts.empty:
        corr_mean = corr_ts.groupby(level=1).mean().reindex(corr_ts.columns, axis=1)

    # ---- correlation full time series wide (pair columns) ----
    def _corr_ts_to_pair_wide(corr_ts: pd.DataFrame) -> pd.DataFrame:
        if corr_ts is None or corr_ts.empty:
            return pd.DataFrame()
        df = corr_ts.copy()
        df.index = df.index.set_names(["date", "factor1"])
        long = (
            df.stack()
            .rename("corr")
            .reset_index()
            .rename(columns={"level_2": "factor2"})
        )
        long = long[long["factor1"] != long["factor2"]].copy()
        a = long["factor1"].astype(str)
        b = long["factor2"].astype(str)
        long["pair"] = np.where(a < b, a + " ~ " + b, b + " ~ " + a)
        long = long.drop(columns=["factor1", "factor2"]).drop_duplicates(
            subset=["date", "pair"]
        )
        wide = long.pivot(index="date", columns="pair", values="corr").sort_index()
        return wide

    corr_pair_wide = _corr_ts_to_pair_wide(corr_ts)

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        wb = writer.book

        # 1) summary
        summary_out.to_excel(writer, sheet_name="summary", index=False)

        # 1.1) summary: 分组均值 bar chart（如果存在 *_mean 列）
        ws = wb["summary"]
        mean_cols = [c for c in summary_out.columns if str(c).endswith(")_mean")]
        if mean_cols:
            # chart data range
            header_row = 1
            first_col = (
                summary_out.columns.get_loc(mean_cols[0]) + 1
            )  # excel is 1-based
            last_col = summary_out.columns.get_loc(mean_cols[-1]) + 1

            bar = BarChart()
            bar.width = 30
            bar.height = 15
            bar.type = "col"
            bar.title = "Average Return For Each Group (means)"
            bar.y_axis.title = "Return"
            bar.x_axis.title = "Row"

            data = Reference(
                ws,
                min_col=first_col,
                max_col=last_col,
                min_row=header_row,
                max_row=ws.max_row,
            )
            cats = Reference(
                ws, min_col=1, min_row=2, max_row=ws.max_row
            )  # index or first column
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.gapWidth = 50
            ws.add_chart(bar, f"A{ws.max_row + 3}")

        # 2) correlation mean matrix sheet
        if corr_mean is not None and not corr_mean.empty:
            corr_mean.to_excel(writer, sheet_name="correlation_mean")

            ws = wb["correlation_mean"]
            rule = ColorScaleRule(
                start_type="num",
                start_value=-1,
                start_color="2F5597",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="num",
                end_value=1,
                end_color="C00000",
            )
            min_row, min_col = 2, 2
            max_row, max_col = ws.max_row, ws.max_column
            cell_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            ws.conditional_formatting.add(cell_range, rule)
            ws.freeze_panes = "B2"

        # 3) correlation full time series wide + boxplot
        if corr_pair_wide is not None and not corr_pair_wide.empty:
            corr_pair_wide.to_excel(writer, sheet_name="correlation_ts")

            # 尝试添加箱线图（openpyxl版本若不支持则跳过）
            try:
                from openpyxl.chart import BoxWhiskerChart

                ws = wb["correlation_ts"]
                chart = BoxWhiskerChart()
                chart.title = "Correlation (pairwise) Boxplot"
                chart.style = 2
                chart.y_axis.title = "Correlation"

                # 数据区域：从第2列开始（第1列是date index）
                max_row = ws.max_row
                max_col = ws.max_column
                data = Reference(
                    ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row
                )
                chart.add_data(data, titles_from_data=True)
                ws.add_chart(chart, f"A{max_row + 3}")
            except Exception:
                pass

        # 4) per test: keep your original exports (IC/group_value/performance)
        for i, (test_key, res) in enumerate(results.items(), start=1):
            if test_key == "correlation":
                continue

            rr = res["raw"]
            ri = res["inc"]

            # IC data
            ic_raw = pd.concat(
                [rr["ic_raw"]] + ([] if ri is None else [ri["ic_raw"]]), axis=1
            )
            ic_raw = pd.concat([ic_raw, ic_raw.cumsum().add_suffix("_cumsum")], axis=1)
            ic_raw.to_excel(writer, sheet_name=f"info_coef_{i}")

            # IC chart
            factor_cols = ic_raw.columns[
                ~ic_raw.columns.astype(str).str.endswith("_cumsum")
            ]
            ws = wb[f"info_coef_{i}"]
            for ii, col in enumerate(factor_cols):
                bar = BarChart()
                bar.width = 30
                bar.height = 20
                bar.type = "col"
                bar.title = f"Information Coefficient {col}"
                bar.y_axis.title = "IC"
                bar.x_axis.title = "Date"

                raw_data = Reference(
                    ws, min_col=2 + ii, max_col=2 + ii, min_row=1, max_row=ws.max_row
                )
                cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                bar.add_data(raw_data, titles_from_data=True)
                bar.set_categories(cats)
                bar.gapWidth = 50

                line = LineChart()
                cum_data = Reference(
                    ws,
                    min_col=2 + len(factor_cols) + ii,
                    max_col=2 + len(factor_cols) + ii,
                    min_row=1,
                    max_row=ws.max_row,
                )
                line.add_data(cum_data, titles_from_data=True)
                for s in line.series:
                    s.graphicalProperties.line = LineProperties(w=int(1 * 12700))
                line.set_categories(cats)
                line.y_axis.axId = 200
                line.y_axis.crosses = "max"
                line.y_axis.title = "Cumulative IC"
                bar += line
                ws.add_chart(bar, f"B{2 + 50*ii}")

            # Group values
            group_values = pd.concat(
                [rr["group_values"]] + ([] if ri is None else [ri["group_values"]]),
                axis=1,
            )
            group_values.to_excel(writer, sheet_name=f"group_value_{i}")

            # Group performance
            group_performance = pd.concat(
                [rr["group_performance"]]
                + ([] if ri is None else [ri["group_performance"]]),
                axis=1,
            )
            group_performance.to_excel(writer, sheet_name=f"group_performance_{i}")


if __name__ == "__main__":
    FACTOR_DATA_PATH = os.getenv("FACTOR_DATA_PATH")
    if not FACTOR_DATA_PATH:
        raise EnvironmentError(
            "Missing env vars: DATASET_PATH and/or FACTOR_DATA_PATH. "
            "Please set them (e.g. in .env) before running."
        )
    factor_source = DuckPQSource(Path(FACTOR_DATA_PATH))
    factor_source.register("quotes_day")
    factor_source.register("instruments_info")

    params = [
        BacktestParams(
            factor_paths=f"example_factor/{name}",  # CONFIG: Placeholder for factor paths
            baseline_factors=[
                "barra/size",
                "barra/momentum",
                "barra/volatility",
                "barra/liquidity",
                "barra/value",
                "barra/profitability",
                "barra/leverage",
                "barra/growth",
            ],
            begin="2015-01-01",
            end="2025-06-30",
        )
        for name in ["example_factor_name"]
    ]

    output_path = (
        "highfreq_skew_kurt__batch2.xlsx"  # CONFIG: Placeholder for output path
    )
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    results = run(factor_source, params)
    save(output_path, results)
